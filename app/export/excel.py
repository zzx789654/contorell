"""比對結果匯出 — 對應 FR-12、AC-12。

**安全考量（CWE-1236 公式注入）**：來源資料（AD 的 displayName、外部系統的欄位）
可能以 ``=`` ``+`` ``-`` ``@`` 開頭。Excel 開啟時會把這類儲存格當公式執行，
可能導致資料外洩或指令執行。因此**所有寫入的字串一律經過消毒**。
"""

from __future__ import annotations

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.comparison.engine import ComparisonResult, MatchStatus, RiskLevel
from app.providers.file_provider import sanitize_for_export

STATUS_LABELS = {
    MatchStatus.ONLY_IN_A: "僅存在於 A",
    MatchStatus.ONLY_IN_B: "僅存在於 B",
    MatchStatus.MATCHED: "相符",
    MatchStatus.ATTRIBUTE_MISMATCH: "屬性不一致",
}

RISK_LABELS = {
    RiskLevel.HIGH: "高風險",
    RiskLevel.MEDIUM: "中風險",
    RiskLevel.LOW: "低風險",
    RiskLevel.NONE: "無風險",
}

RISK_FILLS = {
    RiskLevel.HIGH: PatternFill("solid", fgColor="FFC7CE"),
    RiskLevel.MEDIUM: PatternFill("solid", fgColor="FFEB9C"),
    RiskLevel.LOW: PatternFill("solid", fgColor="DDEBF7"),
}

BASE_HEADERS = [
    "帳號",
    "顯示名稱",
    "比對狀態",
    "風險等級",
    "風險說明",
    "A 來源狀態",
    "B 來源狀態",
    "差異細節",
]


def _status_text(account) -> str:  # type: ignore[no-untyped-def]
    if account is None:
        return "（不存在）"
    return {"enabled": "啟用", "disabled": "停用", "unknown": "未知"}.get(
        account.status.value, "未知"
    )


def _row_values(row, custom_fields: dict[str, dict[str, str]]) -> list[str]:  # type: ignore[no-untyped-def]
    values = [
        row.display_identifier,
        row.display_name,
        STATUS_LABELS[row.status],
        RISK_LABELS[row.risk],
        row.risk_reason,
        _status_text(row.account_a),
        _status_text(row.account_b),
        "；".join(row.differences),
    ]
    annotations = custom_fields.get(row.key, {})
    values.extend(annotations.values())
    return [sanitize_for_export(str(v)) for v in values]


def export_to_excel(
    result: ComparisonResult,
    *,
    custom_field_labels: list[str] | None = None,
    custom_field_values: dict[str, dict[str, str]] | None = None,
) -> bytes:
    """把比對結果匯出成 xlsx。

    Args:
        result: 比對結果。
        custom_field_labels: 自訂欄位的顯示名稱，依顯示順序。
        custom_field_values: {row_key: {field_key: value}} 的標註資料。

    Returns:
        xlsx 檔案的位元組內容。
    """
    labels = custom_field_labels or []
    values = custom_field_values or {}

    workbook = Workbook()

    # --- 摘要頁 ---
    summary_sheet = workbook.active
    summary_sheet.title = "摘要"
    _write_summary(summary_sheet, result)

    # --- 明細頁 ---
    detail = workbook.create_sheet("比對明細")
    headers = [*BASE_HEADERS, *labels]

    detail.append(headers)
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(vertical="center")

    for row in result.rows:
        detail.append(_row_values(row, values))
        fill = RISK_FILLS.get(row.risk)
        if fill is not None:
            detail.cell(row=detail.max_row, column=4).fill = fill

    _autosize_columns(detail, headers)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{detail.max_row}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_summary(sheet, result: ComparisonResult) -> None:  # type: ignore[no-untyped-def]
    summary = result.summary
    title_font = Font(bold=True, size=14)

    sheet["A1"] = "權限比對報告"
    sheet["A1"].font = title_font

    rows = [
        ("", ""),
        ("來源 A", summary.source_a_label),
        ("來源 B", summary.source_b_label),
        ("比對時間", summary.compared_at.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("", ""),
        ("來源 A 帳號數", summary.source_a_count),
        ("來源 B 帳號數", summary.source_b_count),
        ("比對結果總筆數", summary.total_rows),
        ("", ""),
        ("── 比對狀態分佈 ──", ""),
    ]
    rows.extend(
        (STATUS_LABELS[status], summary.status_counts.get(status, 0)) for status in MatchStatus
    )
    rows.append(("", ""))
    rows.append(("── 風險分佈 ──", ""))
    rows.extend((RISK_LABELS[risk], summary.risk_counts.get(risk, 0)) for risk in RiskLevel)

    if summary.warnings:
        rows.append(("", ""))
        rows.append(("── 資料完整性警告 ──", ""))
        rows.extend(("", warning) for warning in summary.warnings)

    for label, value in rows:
        sheet.append([sanitize_for_export(str(label)), sanitize_for_export(str(value))])

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 70

    for row_cells in sheet.iter_rows(min_col=2, max_col=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize_columns(sheet, headers: list[str]) -> None:  # type: ignore[no-untyped-def]
    """依內容長度調整欄寬，並設上限避免超寬欄位。"""
    for index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for cell in sheet[get_column_letter(index)]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 3, 60)


def export_to_csv(
    result: ComparisonResult,
    *,
    custom_field_labels: list[str] | None = None,
    custom_field_values: dict[str, dict[str, str]] | None = None,
) -> bytes:
    """匯出成 CSV。

    以 **UTF-8 BOM** 編碼——Excel 開啟無 BOM 的 UTF-8 CSV 時中文會亂碼（AC-12）。
    """
    labels = custom_field_labels or []
    values = custom_field_values or {}

    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL)

    writer.writerow([*BASE_HEADERS, *labels])
    for row in result.rows:
        writer.writerow(_row_values(row, values))

    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def build_filename(result: ComparisonResult, extension: str) -> str:
    """產生安全的下載檔名。

    只保留安全字元——來源名稱可能含路徑分隔符或引號，
    直接放進 Content-Disposition 會造成標頭注入或路徑穿越。
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw = f"權限比對_{result.summary.source_a_label}_vs_{result.summary.source_b_label}"
    safe = "".join(char if char.isalnum() or char in "_-" else "_" for char in raw)
    return f"{safe[:80]}_{timestamp}.{extension}"
