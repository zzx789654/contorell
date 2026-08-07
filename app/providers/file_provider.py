"""Excel / CSV 帳號來源 — 對應 FR-05、AC-05。

支援 .xlsx 與 .csv，自動偵測編碼（UTF-8 / UTF-8-BOM / Big5）——
台灣企業匯出的 CSV 常是 Big5，若不處理會整份亂碼。

**安全考量**：上傳檔案是外部輸入，一律限制大小與列數（防 zip bomb 與記憶體耗盡），
並以 ``read_only`` 模式開啟 xlsx（不執行任何巨集或外部參照）。
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from charset_normalizer import from_bytes
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.providers.base import (
    Account,
    AccountProvider,
    AccountStatus,
    ConfigurationError,
    DataError,
    FetchResult,
)

logger = logging.getLogger(__name__)

DEFAULT_MAX_BYTES = 10 * 1024 * 1024  # 10MB
DEFAULT_MAX_ROWS = 50_000
PREVIEW_ROWS = 10

# 以 = + - @ 開頭的儲存格在 Excel 中會被當公式執行（CSV injection / CWE-1236）。
# 本系統只讀取不寫回，但匯出時仍須防護，故在此統一定義。
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


@dataclass(slots=True)
class ColumnMapping:
    """欄位名稱到 Account 屬性的對應。"""

    identifier: str
    display_name: str = ""
    email: str = ""
    status: str = ""
    enabled_values: tuple[str, ...] = (
        "true",
        "1",
        "active",
        "enabled",
        "y",
        "yes",
        "啟用",
        "是",
        "有效",
    )


@dataclass(slots=True)
class FileConfig:
    """檔案來源設定。"""

    column_mapping: ColumnMapping
    max_bytes: int = DEFAULT_MAX_BYTES
    max_rows: int = DEFAULT_MAX_ROWS
    sheet_name: str | None = None

    def validate(self) -> None:
        if not self.column_mapping.identifier:
            raise ConfigurationError(
                "必須指定「帳號」欄位對應",
                remediation="請指定檔案中哪一欄是帳號名稱（比對的鍵）。",
            )


def sniff_columns(
    content: bytes, filename: str, *, sheet_name: str | None = None
) -> tuple[list[str], list[dict[str, str]]]:
    """讀取表頭與前幾列，供 UI 顯示預覽並讓管理者設定欄位對應（AC-05）。

    Returns:
        (欄位名稱清單, 前 10 列的預覽資料)
    """
    rows = _read_rows(content, filename, max_rows=PREVIEW_ROWS, sheet_name=sheet_name)
    if not rows:
        raise DataError(
            "檔案沒有資料列",
            remediation="請確認檔案第一列為欄位名稱，且至少有一列資料。",
        )
    return list(rows[0].keys()), rows


def _read_rows(
    content: bytes, filename: str, *, max_rows: int, sheet_name: str | None = None
) -> list[dict[str, str]]:
    """統一把 xlsx / csv 讀成字典清單。"""
    suffix = Path(filename).suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(content, max_rows=max_rows, sheet_name=sheet_name)
    if suffix in (".csv", ".txt"):
        return _read_csv(content, max_rows=max_rows)

    raise DataError(
        f"不支援的檔案格式：{suffix or '（無副檔名）'}",
        remediation="請上傳 .xlsx 或 .csv 檔案。舊版 .xls 請先另存為 .xlsx。",
    )


def _read_xlsx(
    content: bytes, *, max_rows: int, sheet_name: str | None = None
) -> list[dict[str, str]]:
    """讀取 xlsx。

    ``read_only=True`` 串流讀取，避免大檔一次載入記憶體；
    ``data_only=True`` 只取公式的計算結果，不解析公式本身。
    """
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except InvalidFileException as exc:
        raise DataError(
            "無法開啟 Excel 檔案：格式無效或檔案已損毀",
            remediation="請確認檔案為有效的 .xlsx 格式。舊版 .xls 請先另存為 .xlsx。",
        ) from exc
    except Exception as exc:  # openpyxl 對損毀檔案會拋出各種底層例外
        raise DataError(
            f"無法讀取 Excel 檔案：{type(exc).__name__}",
            remediation="請確認檔案未損毀且未加密。",
        ) from exc

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise DataError(
                    f"找不到工作表「{sheet_name}」。可用工作表：{available}",
                    remediation="請選擇存在的工作表名稱。",
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        if sheet is None:
            raise DataError("Excel 檔案中沒有可讀取的工作表")

        rows_iter = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers = _normalize_headers(header_row)
        if not headers:
            raise DataError(
                "第一列沒有有效的欄位名稱",
                remediation="請確認檔案第一列為欄位標題（如「帳號」「姓名」）。",
            )

        result: list[dict[str, str]] = []
        for row in rows_iter:
            if len(result) >= max_rows:
                break
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # 略過空白列
            result.append(_row_to_dict(headers, row))

        return result
    finally:
        workbook.close()


def _read_csv(content: bytes, *, max_rows: int) -> list[dict[str, str]]:
    """讀取 CSV，自動偵測編碼。"""
    text = _decode_bytes(content)

    # 自動偵測分隔符號（逗號 / 分號 / tab）
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel  # 偵測失敗時退回標準逗號分隔

    # 以 reader（非 DictReader）逐列讀取，讓 CSV 與 xlsx 走同一套
    # 「位置索引 → 欄位名」的處理流程，避免兩種格式行為分歧。
    reader = csv.reader(io.StringIO(text), dialect=dialect)

    try:
        header_row = next(reader)
    except StopIteration:
        raise DataError(
            "CSV 檔案沒有欄位標題列",
            remediation="請確認檔案第一列為欄位名稱。",
        ) from None

    headers = _normalize_headers(tuple(header_row))
    if not headers:
        raise DataError(
            "CSV 第一列沒有有效的欄位名稱",
            remediation="請確認檔案第一列為欄位標題（如「帳號」「姓名」）。",
        )

    result: list[dict[str, str]] = []
    for row in reader:
        if len(result) >= max_rows:
            break
        if all(not str(cell or "").strip() for cell in row):
            continue
        result.append(_row_to_dict(headers, tuple(row)))

    return result


def _decode_bytes(content: bytes) -> str:
    """偵測並解碼位元組。

    優先序：BOM → UTF-8 → **繁中編碼（Big5/CP950）** → 自動偵測 → 寬鬆退回。

    為什麼 Big5 要排在自動偵測之前：統計式的編碼偵測器對短文本
    常把 Big5 誤判為其他雙位元組編碼（如 EUC-KR），解出來是看似合法
    但完全錯誤的韓文字元——這種「錯得很安靜」的結果比解碼失敗更糟，
    因為使用者會拿到一份亂碼名單卻以為比對成功。
    本系統的使用場景是台灣企業，Big5/CP950 是最可能的非 UTF-8 編碼，
    因此先嚴格嘗試（``errors="strict"``），成功才採用。
    """
    if content.startswith(b"\xef\xbb\xbf"):
        return content[3:].decode("utf-8", errors="replace")

    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        pass

    # CP950 是 Big5 的微軟擴充版，涵蓋範圍更廣，故先試 CP950
    for encoding in ("cp950", "big5"):
        try:
            decoded = content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
        logger.info("檔案編碼判定為 %s（繁體中文）", encoding)
        return decoded

    matches = from_bytes(content)
    best = matches.best()
    if best is not None:
        logger.info("檔案編碼自動偵測為：%s", best.encoding)
        return str(best)

    logger.warning("無法可靠偵測檔案編碼，以 UTF-8 寬鬆解碼，可能出現替代字元")
    return content.decode("utf-8", errors="replace")


def _normalize_headers(header_row: tuple) -> dict[int, str]:
    """把表頭正規化成 ``{欄位位置: 清理後的欄位名}``。

    一律以**位置索引**為鍵，不因來源格式而改變鍵型別——
    先前的實作對 CSV 用名稱、對 xlsx 用索引，導致下游取值時
    型別對不上而靜默回傳空資料。單一鍵型別讓契約明確、不易出錯。
    """
    return {
        index: str(cell).strip()
        for index, cell in enumerate(header_row)
        if cell is not None and str(cell).strip()
    }


def _row_to_dict(headers: dict[int, str], row: tuple) -> dict[str, str]:
    """依表頭位置把整列轉成 ``{欄位名: 值}``。"""
    return {
        name: _clean_cell(row[index] if index < len(row) else None)
        for index, name in headers.items()
    }


def _clean_cell(value: object) -> str:
    """把儲存格轉成乾淨的字串。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))  # 避免帳號 12345 變成 "12345.0"
    return str(value).strip()


class FileProvider(AccountProvider):
    """從上傳的 Excel / CSV 檔案讀取帳號清單。"""

    def __init__(
        self, content: bytes, filename: str, config: FileConfig, *, label: str = ""
    ) -> None:
        config.validate()

        if len(content) > config.max_bytes:
            raise DataError(
                f"檔案大小 {len(content) / 1024 / 1024:.1f}MB 超過上限 "
                f"{config.max_bytes / 1024 / 1024:.0f}MB",
                remediation="請分批上傳，或先移除不必要的欄位與工作表。",
            )
        if not content:
            raise DataError("檔案是空的")

        self._content = content
        self._filename = filename
        self._config = config
        self._label = label or filename

    @property
    def label(self) -> str:
        return self._label

    def fetch(self) -> FetchResult:
        rows = _read_rows(
            self._content,
            self._filename,
            max_rows=self._config.max_rows,
            sheet_name=self._config.sheet_name,
        )
        warnings: list[str] = []

        if not rows:
            raise DataError(
                "檔案中沒有資料列",
                remediation="請確認檔案第一列為欄位標題，且下方至少有一列資料。",
            )

        mapping = self._config.column_mapping
        available = set(rows[0].keys())

        if mapping.identifier not in available:
            raise ConfigurationError(
                f"檔案中找不到欄位「{mapping.identifier}」。"
                f"可用欄位：{', '.join(sorted(available))}",
                remediation="請重新設定欄位對應，選擇檔案中實際存在的欄位。",
            )

        accounts: list[Account] = []
        skipped = 0

        for row in rows:
            identifier = row.get(mapping.identifier, "").strip()
            if not identifier:
                skipped += 1
                continue

            status = AccountStatus.UNKNOWN
            if mapping.status and mapping.status in row:
                raw = row[mapping.status].strip().lower()
                if raw:
                    status = (
                        AccountStatus.ENABLED
                        if raw in mapping.enabled_values
                        else AccountStatus.DISABLED
                    )

            accounts.append(
                Account(
                    identifier=identifier,
                    display_name=row.get(mapping.display_name, ""),
                    email=row.get(mapping.email, ""),
                    status=status,
                )
            )

        if skipped:
            warnings.append(f"有 {skipped} 列的「{mapping.identifier}」欄位是空的，已略過。")

        if len(rows) >= self._config.max_rows:
            warnings.append(f"檔案列數已達上限 {self._config.max_rows:,} 列，超出部分未讀取。")

        return FetchResult(
            accounts=accounts,
            fetched_at=datetime.now(UTC),
            source_label=self._label,
            total_reported=len(rows),
            warnings=warnings,
            diagnostics={
                "filename": self._filename,
                "rows_read": str(len(rows)),
                "size_bytes": str(len(self._content)),
            },
        )

    def test_connection(self) -> dict[str, str]:
        """檔案來源無連線概念，回傳解析結果作為驗證。"""
        columns, preview = sniff_columns(
            self._content, self._filename, sheet_name=self._config.sheet_name
        )
        mapped = self._config.column_mapping.identifier
        return {
            "status": "檔案解析成功",
            "columns": ", ".join(columns),
            "preview_rows": str(len(preview)),
            "identifier_column": (
                f"{mapped}（存在）" if mapped in columns else f"{mapped}（**檔案中不存在**）"
            ),
        }


def sanitize_for_export(value: str) -> str:
    """防止 CSV/Excel 公式注入（CWE-1236）。

    匯出時，以 = + - @ 開頭的儲存格會被 Excel 當公式執行。
    來源資料（如 AD 的 displayName）可能含這類字元，
    因此匯出前一律加上單引號前綴使其成為文字。
    """
    if value and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value
