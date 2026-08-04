"""Excel / CSV 來源測試 — 對應 FR-05、AC-05。

重點：Big5 編碼（台灣企業匯出的 CSV 常見）、欄位對應錯誤的可行動提示、
以及惡意/畸形檔案不可造成崩潰。
"""

import io

import pytest
from openpyxl import Workbook

from app.providers.base import AccountStatus, ConfigurationError, DataError
from app.providers.file_provider import (
    ColumnMapping,
    FileConfig,
    FileProvider,
    sniff_columns,
)


def make_xlsx(rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def make_config(identifier: str = "帳號", **kwargs) -> FileConfig:
    return FileConfig(
        column_mapping=ColumnMapping(identifier=identifier, **kwargs)
    )


class TestCsvParsing:
    def test_utf8_csv(self):
        content = "帳號,姓名\nalice,愛麗絲\nbob,鮑伯\n".encode()
        provider = FileProvider(content, "users.csv", make_config(display_name="姓名"))

        result = provider.fetch()

        assert result.count == 2
        assert result.accounts[0].identifier == "alice"
        assert result.accounts[0].display_name == "愛麗絲"

    def test_utf8_bom_csv(self):
        """Excel 匯出的 UTF-8 CSV 帶 BOM。"""
        content = "﻿帳號,姓名\nalice,愛麗絲\n".encode()
        provider = FileProvider(content, "users.csv", make_config(display_name="姓名"))

        result = provider.fetch()

        assert result.count == 1
        assert result.accounts[0].identifier == "alice"

    def test_big5_csv(self):
        """AC-05：台灣企業從舊系統匯出的 CSV 多為 Big5，必須正確解碼。"""
        content = "帳號,姓名\nalice,愛麗絲\nbob,鮑伯\n".encode("big5")
        provider = FileProvider(content, "users.csv", make_config(display_name="姓名"))

        result = provider.fetch()

        assert result.count == 2
        assert result.accounts[0].display_name == "愛麗絲"
        assert result.accounts[1].display_name == "鮑伯"

    def test_semicolon_delimiter_detected(self):
        content = "帳號;姓名\nalice;Alice\n".encode()
        provider = FileProvider(content, "users.csv", make_config(display_name="姓名"))

        assert provider.fetch().count == 1

    def test_blank_rows_skipped(self):
        content = "帳號\nalice\n\n\nbob\n".encode()

        result = FileProvider(content, "u.csv", make_config()).fetch()

        assert result.count == 2

    def test_rows_with_empty_identifier_warned(self):
        content = "帳號,姓名\nalice,A\n,B\nbob,C\n".encode()

        result = FileProvider(content, "u.csv", make_config(display_name="姓名")).fetch()

        assert result.count == 2
        assert any("空的" in w for w in result.warnings)


class TestXlsxParsing:
    def test_basic_xlsx(self):
        content = make_xlsx([["帳號", "姓名"], ["alice", "愛麗絲"], ["bob", "鮑伯"]])

        result = FileProvider(content, "u.xlsx", make_config(display_name="姓名")).fetch()

        assert result.count == 2
        assert result.accounts[1].identifier == "bob"

    def test_numeric_account_not_turned_into_float(self):
        """帳號 12345 不可變成 '12345.0'——那會導致比對永遠對不上。"""
        content = make_xlsx([["帳號"], [12345], [67890]])

        result = FileProvider(content, "u.xlsx", make_config()).fetch()

        assert result.accounts[0].identifier == "12345"
        assert result.accounts[1].identifier == "67890"

    def test_status_column_mapped(self):
        content = make_xlsx(
            [["帳號", "狀態"], ["alice", "啟用"], ["bob", "停用"]]
        )

        result = FileProvider(
            content, "u.xlsx", make_config(status="狀態")
        ).fetch()

        assert result.accounts[0].status is AccountStatus.ENABLED
        assert result.accounts[1].status is AccountStatus.DISABLED

    def test_sheet_selection(self):
        workbook = Workbook()
        workbook.active.title = "第一頁"
        workbook.active.append(["帳號"])
        workbook.active.append(["wrong"])
        second = workbook.create_sheet("目標頁")
        second.append(["帳號"])
        second.append(["correct"])
        buffer = io.BytesIO()
        workbook.save(buffer)

        config = FileConfig(
            column_mapping=ColumnMapping(identifier="帳號"), sheet_name="目標頁"
        )
        result = FileProvider(buffer.getvalue(), "u.xlsx", config).fetch()

        assert result.accounts[0].identifier == "correct"

    def test_missing_sheet_gives_actionable_error(self):
        content = make_xlsx([["帳號"], ["alice"]], sheet_name="實際名稱")
        config = FileConfig(
            column_mapping=ColumnMapping(identifier="帳號"), sheet_name="不存在"
        )

        with pytest.raises(DataError) as exc:
            FileProvider(content, "u.xlsx", config).fetch()

        assert "實際名稱" in exc.value.message  # 列出可用工作表


class TestErrorHandling:
    """AC-05：欄位對應錯誤時要給出可修正的提示，而非籠統錯誤。"""

    def test_missing_column_lists_available_ones(self):
        content = make_xlsx([["使用者", "姓名"], ["alice", "A"]])

        with pytest.raises(ConfigurationError) as exc:
            FileProvider(content, "u.xlsx", make_config("帳號")).fetch()

        assert "使用者" in exc.value.message  # 告訴使用者實際有哪些欄位
        assert exc.value.remediation

    def test_unsupported_extension_rejected(self):
        with pytest.raises(DataError, match="不支援"):
            FileProvider(b"data", "users.pdf", make_config()).fetch()

    def test_empty_file_rejected(self):
        with pytest.raises(DataError, match="空"):
            FileProvider(b"", "u.csv", make_config())

    def test_oversized_file_rejected(self):
        config = FileConfig(column_mapping=ColumnMapping(identifier="帳號"), max_bytes=100)

        with pytest.raises(DataError, match="超過上限"):
            FileProvider(b"x" * 200, "u.csv", config)

    def test_corrupt_xlsx_gives_clear_error(self):
        with pytest.raises(DataError):
            FileProvider(b"not a real xlsx file", "u.xlsx", make_config()).fetch()

    def test_header_only_file_rejected(self):
        content = "帳號,姓名\n".encode()

        with pytest.raises(DataError, match="沒有資料"):
            FileProvider(content, "u.csv", make_config()).fetch()

    def test_row_limit_produces_warning(self):
        rows = [["帳號"]] + [[f"user{i}"] for i in range(20)]
        config = FileConfig(
            column_mapping=ColumnMapping(identifier="帳號"), max_rows=10
        )

        result = FileProvider(make_xlsx(rows), "u.xlsx", config).fetch()

        assert result.count == 10
        assert any("上限" in w for w in result.warnings)


class TestColumnSniffing:
    """AC-05：上傳後顯示預覽，讓管理者設定欄位對應。"""

    def test_returns_columns_and_preview(self):
        content = make_xlsx(
            [["帳號", "姓名", "Email"], ["alice", "A", "a@x.com"], ["bob", "B", "b@x.com"]]
        )

        columns, preview = sniff_columns(content, "u.xlsx")

        assert columns == ["帳號", "姓名", "Email"]
        assert len(preview) == 2
        assert preview[0]["帳號"] == "alice"

    def test_preview_limited_to_ten_rows(self):
        rows = [["帳號"]] + [[f"user{i}"] for i in range(50)]

        _, preview = sniff_columns(make_xlsx(rows), "u.xlsx")

        assert len(preview) == 10

    def test_csv_sniffing(self):
        content = "帳號,部門\nalice,財務\n".encode()

        columns, preview = sniff_columns(content, "u.csv")

        assert "帳號" in columns
        assert "部門" in columns


class TestConnectionTest:
    def test_reports_column_presence(self):
        content = make_xlsx([["帳號", "姓名"], ["alice", "A"]])

        diagnostics = FileProvider(content, "u.xlsx", make_config()).test_connection()

        assert diagnostics["status"] == "檔案解析成功"
        assert "存在" in diagnostics["identifier_column"]

    def test_reports_missing_column(self):
        content = make_xlsx([["使用者"], ["alice"]])

        diagnostics = FileProvider(
            content, "u.xlsx", make_config("帳號")
        ).test_connection()

        assert "不存在" in diagnostics["identifier_column"]
