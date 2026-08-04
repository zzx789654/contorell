"""匯出測試 — 對應 FR-12、AC-12，以及 CSV 公式注入防護（CWE-1236）。"""

import io
from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from app.comparison.engine import compare
from app.export.excel import build_filename, export_to_csv, export_to_excel
from app.providers.base import Account, AccountStatus, FetchResult
from app.providers.file_provider import sanitize_for_export


def make_result(accounts, label="來源"):
    return FetchResult(accounts=accounts, fetched_at=datetime.now(UTC), source_label=label)


@pytest.fixture
def sample_comparison():
    ad = make_result(
        [
            Account(identifier="leaver", display_name="離職者", status=AccountStatus.DISABLED),
            Account(identifier="normal", display_name="正常員工", status=AccountStatus.ENABLED),
        ],
        "AD 財務部",
    )
    erp = make_result(
        [
            Account(identifier="leaver", display_name="離職者", status=AccountStatus.ENABLED),
            Account(identifier="normal", display_name="正常員工", status=AccountStatus.ENABLED),
            Account(identifier="ghost", display_name="孤兒帳號", status=AccountStatus.ENABLED),
        ],
        "ERP 系統",
    )
    return compare(ad, erp, a_is_authoritative=True)


class TestFormulaInjectionDefense:
    """CWE-1236：匯出的儲存格不可被 Excel 當公式執行。"""

    @pytest.mark.parametrize(
        "payload",
        [
            "=1+1",
            "=cmd|'/c calc'!A1",
            "+1234",
            "-1234",
            "@SUM(A1:A9)",
            "\tTabbed",
        ],
    )
    def test_dangerous_prefixes_neutralized(self, payload):
        result = sanitize_for_export(payload)

        assert result.startswith("'"), f"未消毒的危險字串：{payload}"
        assert result == "'" + payload

    @pytest.mark.parametrize("safe", ["normal.user", "王小明", "user@example.com", "", "123abc"])
    def test_safe_values_unchanged(self, safe):
        assert sanitize_for_export(safe) == safe

    def test_malicious_display_name_sanitized_in_export(self, sample_comparison):
        """實務情境：AD 的 displayName 含公式字元。"""
        ad = make_result(
            [Account(identifier="evil", display_name="=cmd|'/c calc'!A1")], "AD"
        )
        erp = make_result([], "ERP")
        result = compare(ad, erp)

        content = export_to_csv(result)
        text = content.decode("utf-8-sig")

        # 公式必須已被單引號前綴中和
        assert "\"'=cmd" in text or "'=cmd" in text
        assert '"=cmd' not in text


class TestExcelExport:
    def test_produces_valid_xlsx(self, sample_comparison):
        content = export_to_excel(sample_comparison)

        workbook = load_workbook(io.BytesIO(content))

        assert "摘要" in workbook.sheetnames
        assert "比對明細" in workbook.sheetnames

    def test_detail_sheet_contains_all_rows(self, sample_comparison):
        content = export_to_excel(sample_comparison)
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook["比對明細"]

        # 表頭 + 3 筆資料
        assert sheet.max_row == 4

    def test_summary_includes_source_counts(self, sample_comparison):
        """來源筆數必須出現在報表中，供人工交叉檢查（風險 R-02）。"""
        content = export_to_excel(sample_comparison)
        workbook = load_workbook(io.BytesIO(content))
        values = [
            str(value)
            for row in workbook["摘要"].iter_rows(values_only=True)
            for value in row
            if value
        ]

        assert any("來源 A 帳號數" in v for v in values)
        assert any("AD 財務部" in v for v in values)

    def test_custom_fields_appended(self, sample_comparison):
        content = export_to_excel(
            sample_comparison,
            custom_field_labels=["處理狀態", "負責人"],
            custom_field_values={"leaver": {"處理狀態": "已通知", "負責人": "王經理"}},
        )
        workbook = load_workbook(io.BytesIO(content))
        headers = [cell.value for cell in workbook["比對明細"][1]]

        assert "處理狀態" in headers
        assert "負責人" in headers

    def test_high_risk_rows_present(self, sample_comparison):
        content = export_to_excel(sample_comparison)
        workbook = load_workbook(io.BytesIO(content))
        rows = list(workbook["比對明細"].iter_rows(min_row=2, values_only=True))

        risk_column = [row[3] for row in rows]
        assert "高風險" in risk_column


class TestCsvExport:
    def test_utf8_bom_for_excel_compatibility(self, sample_comparison):
        """AC-12：中文不可亂碼。Excel 需要 BOM 才能正確識別 UTF-8。"""
        content = export_to_csv(sample_comparison)

        assert content.startswith(b"\xef\xbb\xbf")

    def test_chinese_content_readable(self, sample_comparison):
        content = export_to_csv(sample_comparison)
        text = content.decode("utf-8-sig")

        assert "離職者" in text
        assert "高風險" in text

    def test_all_rows_exported(self, sample_comparison):
        content = export_to_csv(sample_comparison)
        lines = content.decode("utf-8-sig").strip().splitlines()

        assert len(lines) == 4  # 表頭 + 3 筆


class TestFilename:
    def test_removes_unsafe_characters(self, sample_comparison):
        """檔名不可含路徑分隔符或引號，避免標頭注入與路徑穿越。"""
        filename = build_filename(sample_comparison, "xlsx")

        for char in ('/', '\\', '"', "'", "\n", "\r", ".."):
            assert char not in filename.replace(".xlsx", "")

    def test_has_correct_extension(self, sample_comparison):
        assert build_filename(sample_comparison, "csv").endswith(".csv")
        assert build_filename(sample_comparison, "xlsx").endswith(".xlsx")
